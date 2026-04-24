#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import glob
import json
import math
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any


def to_float(x: Any) -> float | None:
    if x is None or x == "":
        return None
    try:
        v = float(x)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except Exception:
        return None


def mean(values: list[float | None]) -> float | None:
    vals = [v for v in values if v is not None]
    return sum(vals) / len(vals) if vals else None


def median(values: list[float | None]) -> float | None:
    vals = [v for v in values if v is not None]
    return statistics.median(vals) if vals else None


def safe_round(x: Any, ndigits: int = 4) -> Any:
    v = to_float(x)
    return round(v, ndigits) if v is not None else ""


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
                row["source_file"] = str(path)
                rows.append(row)
            except json.JSONDecodeError as e:
                print(f"Skipping invalid JSON in {path}:{line_no}: {e}")
    return rows


def load_rows(results_dir: Path, pattern: str) -> list[dict[str, Any]]:
    files = sorted(Path(p) for p in glob.glob(str(results_dir / pattern)))
    files = [p for p in files if not p.name.endswith("_summary.jsonl")]
    rows: list[dict[str, Any]] = []
    for p in files:
        rows.extend(read_jsonl(p))
    return rows


def normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        latency = to_float(r.get("latency_s"))
        test_time = to_float(r.get("test_time_s"))
        completion_tokens = to_float(r.get("completion_tokens"))
        raw_generation = r.get("raw_generation") or ""
        extracted_code = r.get("extracted_code") or ""

        if r.get("completion_tokens_per_second") is not None:
            tps = to_float(r.get("completion_tokens_per_second"))
        elif completion_tokens is not None and latency and latency > 0:
            tps = completion_tokens / latency
        else:
            tps = None

        chars_per_second = to_float(r.get("generated_chars_per_second"))
        if chars_per_second is None and latency and latency > 0:
            chars_per_second = len(raw_generation) / latency

        nr = dict(r)
        nr["latency_s"] = latency
        nr["test_time_s"] = test_time
        nr["total_time_s"] = (latency or 0.0) + (test_time or 0.0)
        nr["prompt_tokens"] = to_float(r.get("prompt_tokens"))
        nr["completion_tokens"] = completion_tokens
        nr["total_tokens"] = to_float(r.get("total_tokens"))
        nr["completion_tokens_per_second"] = tps
        nr["generated_chars"] = len(raw_generation)
        nr["code_chars"] = len(extracted_code)
        nr["generated_chars_per_second"] = chars_per_second
        nr["passed"] = bool(r.get("passed"))
        out.append(nr)
    return out


def summarize_by_model(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        groups[str(r.get("model", "unknown"))].append(r)

    summary = []
    for model, rs in sorted(groups.items()):
        attempts = len(rs)
        passed = sum(1 for r in rs if r["passed"])
        task_ids = sorted({str(r.get("task_id", "")) for r in rs})
        summary.append({
            "model": model,
            "tasks": len(task_ids),
            "attempts": attempts,
            "passed": passed,
            "failed": attempts - passed,
            "pass_rate_pct": safe_round(100.0 * passed / attempts if attempts else None, 2),
            "avg_generation_latency_s": safe_round(mean([r.get("latency_s") for r in rs]), 4),
            "median_generation_latency_s": safe_round(median([r.get("latency_s") for r in rs]), 4),
            "total_generation_latency_s": safe_round(sum((r.get("latency_s") or 0.0) for r in rs), 4),
            "avg_test_time_s": safe_round(mean([r.get("test_time_s") for r in rs]), 4),
            "total_test_time_s": safe_round(sum((r.get("test_time_s") or 0.0) for r in rs), 4),
            "avg_total_time_s": safe_round(mean([r.get("total_time_s") for r in rs]), 4),
            "total_eval_time_s": safe_round(sum((r.get("total_time_s") or 0.0) for r in rs), 4),
            "avg_prompt_tokens": safe_round(mean([r.get("prompt_tokens") for r in rs]), 2),
            "avg_completion_tokens": safe_round(mean([r.get("completion_tokens") for r in rs]), 2),
            "total_completion_tokens": safe_round(sum((r.get("completion_tokens") or 0.0) for r in rs), 0),
            "avg_completion_tokens_per_second": safe_round(mean([r.get("completion_tokens_per_second") for r in rs]), 4),
            "avg_generated_chars": safe_round(mean([r.get("generated_chars") for r in rs]), 2),
            "avg_generated_chars_per_second": safe_round(mean([r.get("generated_chars_per_second") for r in rs]), 4),
        })
    return summary


def summarize_by_task(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        groups[(str(r.get("model", "unknown")), str(r.get("task_id", "unknown")))].append(r)

    details = []
    for (model, task_id), rs in sorted(groups.items()):
        attempts = len(rs)
        passed = sum(1 for r in rs if r["passed"])
        last = rs[-1]
        details.append({
            "model": model,
            "task_id": task_id,
            "attempts": attempts,
            "passed": passed,
            "failed": attempts - passed,
            "pass_rate_pct": safe_round(100.0 * passed / attempts if attempts else None, 2),
            "avg_generation_latency_s": safe_round(mean([r.get("latency_s") for r in rs]), 4),
            "best_generation_latency_s": safe_round(min([r.get("latency_s") for r in rs if r.get("latency_s") is not None], default=None), 4),
            "avg_test_time_s": safe_round(mean([r.get("test_time_s") for r in rs]), 4),
            "avg_completion_tokens": safe_round(mean([r.get("completion_tokens") for r in rs]), 2),
            "avg_completion_tokens_per_second": safe_round(mean([r.get("completion_tokens_per_second") for r in rs]), 4),
            "avg_generated_chars_per_second": safe_round(mean([r.get("generated_chars_per_second") for r in rs]), 4),
            "last_returncode": last.get("returncode", ""),
            "last_error_short": str(last.get("stderr", ""))[:300].replace("\n", " | "),
        })
    return details


def make_task_comparison(task_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_task: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    models = sorted({r["model"] for r in task_rows})
    for r in task_rows:
        by_task[r["task_id"]][r["model"]] = r

    rows = []
    for task_id in sorted(by_task):
        row: dict[str, Any] = {"task_id": task_id}
        for model in models:
            r = by_task[task_id].get(model, {})
            row[f"{model}_pass_rate_pct"] = r.get("pass_rate_pct", "")
            row[f"{model}_avg_latency_s"] = r.get("avg_generation_latency_s", "")
            row[f"{model}_avg_tps"] = r.get("avg_completion_tokens_per_second", "")
            row[f"{model}_avg_chars_per_s"] = r.get("avg_generated_chars_per_second", "")
        rows.append(row)
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields: list[str] = []
    for r in rows:
        for k in r.keys():
            if k not in fields:
                fields.append(k)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        if not rows:
            ws.append(["No data"])
            continue
        fields: list[str] = []
        for r in rows:
            for k in r.keys():
                if k not in fields:
                    fields.append(k)
        ws.append(fields)
        for r in rows:
            ws.append([r.get(k, "") for k in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=False)
        for col_idx, field in enumerate(fields, start=1):
            values = [str(field)] + [str(r.get(field, "")) for r in rows[:200]]
            width = min(max(max(len(v) for v in values) + 2, 10), 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)
    return True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--results-dir", default="results")
    parser.add_argument("--pattern", default="*.jsonl")
    parser.add_argument("--out-prefix", default="results/qwen_report")
    args = parser.parse_args()

    results_dir = Path(args.results_dir)
    out_prefix = Path(args.out_prefix)

    rows = normalize_rows(load_rows(results_dir, args.pattern))
    if not rows:
        raise SystemExit(f"No JSONL rows found in {results_dir}/{args.pattern}")

    model_summary = summarize_by_model(rows)
    task_details = summarize_by_task(rows)
    task_comparison = make_task_comparison(task_details)

    files = {
        "model_summary": out_prefix.with_name(out_prefix.name + "_model_summary.csv"),
        "task_details": out_prefix.with_name(out_prefix.name + "_task_details.csv"),
        "task_comparison": out_prefix.with_name(out_prefix.name + "_task_comparison.csv"),
    }

    write_csv(files["model_summary"], model_summary)
    write_csv(files["task_details"], task_details)
    write_csv(files["task_comparison"], task_comparison)

    xlsx_path = out_prefix.with_suffix(".xlsx")
    xlsx_ok = write_xlsx(xlsx_path, {
        "Model Summary": model_summary,
        "Task Details": task_details,
        "Task Comparison": task_comparison,
        "Raw Results": rows,
    })

    print("Report generated:")
    for p in files.values():
        print(f"  {p}")
    if xlsx_ok:
        print(f"  {xlsx_path}")
    else:
        print("  Excel not written because openpyxl is not installed.")
        print("  Install with: uv pip install openpyxl")

    print("\nModel summary:")
    for r in model_summary:
        print(
            f"  {r['model']}: pass_rate={r['pass_rate_pct']}%, "
            f"avg_latency={r['avg_generation_latency_s']}s, "
            f"avg_tps={r['avg_completion_tokens_per_second'] or 'N/A'}, "
            f"avg_chars/s={r['avg_generated_chars_per_second']}"
        )


if __name__ == "__main__":
    main()
